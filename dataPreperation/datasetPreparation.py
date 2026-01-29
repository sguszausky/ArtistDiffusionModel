from PIL import Image
from pathlib import Path
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import re


validEndings = {".jpg", ".jpeg", ".png"}


def isChecked(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    return s in {"true", "wahr", "1", "x", "yes", "ja", "checked"}


def safeFilename(stem: str) -> str:
    stem = stem.strip()
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", stem)
    stem = stem.rstrip(". ")
    return stem or "unnamed"


def excelTowsToTxt(
    xlsx_path: str | Path,
    sheet_name: str | None = None,
    image_col: str = "image_name",
    base_col: str = "caption_base",
    sep: str = ", ",
) -> int:
    xlsx_path = Path(xlsx_path)
    out_dir = xlsx_path.parent / "captions"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read header row (row 1)
    headers = [c.value for c in ws[1]]
    if image_col not in headers or base_col not in headers:
        raise ValueError(
            f"Missing required columns. Found: {headers}\n"
            f"Need: '{image_col}' and '{base_col}'"
        )

    col_index = {name: i for i, name in enumerate(headers)}  # 0-based
    image_i = col_index[image_col]
    base_i = col_index[base_col]

    checkbox_cols = [h for h in headers if h not in (image_col, base_col)]

    written = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        image_name = row[image_i]
        base = row[base_i]

        if image_name is None or str(image_name).strip() == "":
            continue

        stem = Path(str(image_name)).stem
        txt_name = safeFilename(stem) + ".txt"

        checked_headers = []
        for h in checkbox_cols:
            v = row[col_index[h]] if col_index[h] < len(row) else None
            if isChecked(v):
                checked_headers.append(str(h))

        base_str = "" if base is None else str(base).strip()
        caption = base_str
        if checked_headers:
            caption = (caption + sep if caption else "") + sep.join(checked_headers)

        (out_dir / txt_name).write_text(caption, encoding="utf-8")
        written += 1

    return written

def imageTooSmall(img_path: Path, minWidth: int, minHeight: int) -> tuple[bool, int, int]:
    with Image.open(img_path) as img:
        img.load()
        w, h = img.size
    return (w < minWidth or h < minHeight), w, h


def filterBySize(imageDir: str,minWidth: int,minHeight: int,):

    imageDir = Path(imageDir)
    to_delete = []

    for img_path in imageDir.iterdir():
        if img_path.suffix.lower() not in validEndings:
            continue
        try:
            too_small, w, h = imageTooSmall(img_path, minWidth, minHeight)
            if too_small:
                to_delete.append({"path": img_path, "width": w, "height": h})
        except Exception as e:
            print(f"Error occured at {img_path.name}: {e}")

    print(f"\nFound {len(to_delete)} images smaller as {minWidth}×{minHeight}")
    for item in to_delete:
        path = item["path"]
        width = item["width"]
        height = item["height"]
        print(f"{path.name}: {width}×{height}")
    return to_delete



def deleteImages(to_delete: list[dict],retries: int = 3,wait: float = 0.5,):
    deleted = 0
    locked = 0

    for item in to_delete:
        img_path: Path = item["path"]
        w = item["width"]
        h = item["height"]

        if not img_path.exists():
            continue

        for attempt in range(retries):
            try:
                img_path.unlink()
                deleted += 1
                print(f"Deleted: {img_path.name} ({w}×{h})")
                break
            except PermissionError:
                if attempt < retries - 1:
                    time.sleep(wait)
                else:
                    locked += 1
                    print(f"Could not delete: {img_path.name}, skipped")

    print("\nResult")
    print("Deleted:", deleted)
    print("Skipped:", locked)

def moveImages(
    to_move: list[dict],
    dest_folder: str | Path,
    retries: int = 3,
    wait: float = 0.5,
):
    moved = 0
    locked = 0
    missing = 0

    dest_folder = Path(dest_folder).expanduser().resolve()
    dest_folder.mkdir(parents=True, exist_ok=True)

    for item in to_move:
        img_path: Path = item["path"]
        w = item.get("width")
        h = item.get("height")

        if not img_path.exists():
            missing += 1
            continue

        target_path = dest_folder / img_path.name

        # Avoid overwriting if a file with same name already exists
        if target_path.exists():
            stem, suffix = img_path.stem, img_path.suffix
            k = 1
            while True:
                candidate = dest_folder / f"{stem}_{k}{suffix}"
                if not candidate.exists():
                    target_path = candidate
                    break
                k += 1

        for attempt in range(retries):
            try:
                shutil.move(str(img_path), str(target_path))
                moved += 1
                if w and h:
                    print(f"Moved: {img_path.name} -> {target_path.name} ({w}×{h})")
                else:
                    print(f"Moved: {img_path.name} -> {target_path.name}")
                break
            except PermissionError:
                if attempt < retries - 1:
                    time.sleep(wait)
                else:
                    locked += 1
                    print(f"Could not move: {img_path.name}, skipped (file locked?)")

    print("\nResult")
    print("Moved:", moved)
    print("Skipped (locked):", locked)
    print("Missing:", missing)


def cropBottomPercent(imageDir: str,outDir: str,cropPercent: float,overwrite: bool = False,quality: int = 95):
    """
    Crops a fixed percentage from the bottom of every image.

    Args:
        imageDir: input folder
        outDir: output folder (ignored if overwrite=True; then writes into imageDir)
        cropPercent: e.g. 12.0 means cut off 12% from bottom
        overwrite: if True, overwrite originals (NOT recommended)
        quality: JPEG quality
    """
    in_dir = Path(imageDir)
    out_dir = in_dir if overwrite else Path(outDir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if cropPercent <= 0 or cropPercent >= 100:
        raise ValueError("cropPercent must be between 0 and 100 (exclusive).")

    processed = 0
    skipped = 0
    failed = 0

    for img_path in in_dir.iterdir():
        if img_path.suffix.lower() not in validEndings:
            continue

        try:
            with Image.open(img_path) as img:
                img.load()
                img = img.convert("RGB")
                w, h = img.size

                cut_px = int(h * (cropPercent / 100.0))
                new_h = h - cut_px

                if new_h < 1:
                    skipped += 1
                    print(f"Skipped (too small after crop): {img_path.name}")
                    continue

                cropped = img.crop((0, 0, w, new_h))
                out_path = out_dir / img_path.name
                cropped.save(out_path, quality=quality)

            processed += 1
            if processed <= 3 or processed % 25 == 0:
                print(f"Cropped: {img_path.name} ({w}×{h} -> {w}×{new_h})")

        except Exception as e:
            failed += 1
            print(f"Error occured at {img_path.name}: {e}")

    print("\nResult")
    print("Processed:", processed)
    print("Skipped:", skipped)
    print("Failed:", failed)
    print("Output:", out_dir)


def numberImagesandCreateDict(folder: str, excel_name: str = "mapping.xlsx", start_index: int = 1):
    folder_path = Path(folder).expanduser().resolve()
    if not folder_path.is_dir():
        raise NotADirectoryError(f"Folder not found: {folder_path}")

    images = sorted(
        [p for p in folder_path.iterdir() if p.is_file() and p.suffix.lower() in validEndings],
        key=lambda p: p.name.lower()
    )

    if not images:
        print("No images found.")
        return None

    targets = []
    idx = start_index
    for p in images:
        new_name = f"img_{idx}{p.suffix.lower()}"   # <-- changed here
        targets.append((p, folder_path / new_name))
        idx += 1

    temp_pairs = []
    for i, (src, _) in enumerate(targets, start=1):
        tmp = src.with_name(f"__tmp_rename__{i}__{src.name}")
        temp_pairs.append((src, tmp))

    for src, tmp in temp_pairs:
        os.rename(src, tmp)

    mapping_rows = []
    for (orig_src, final_dst), (_, tmp_src) in zip(targets, temp_pairs):
        os.rename(tmp_src, final_dst)
        mapping_rows.append((orig_src.name, final_dst.name))

    wb = Workbook()
    ws = wb.active
    ws.title = "mapping"
    ws.append(["original_name", "new_name"])
    for row in mapping_rows:
        ws.append(list(row))

    excel_out = folder_path / excel_name
    wb.save(excel_out)

    print(f"Renamed {len(mapping_rows)} images.")
    print(f"Mapping saved to: {excel_out}")
    return excel_out

